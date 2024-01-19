from urllib.request import urlopen
import asyncio
import re
import requests
from PIL import Image
from openpyxl import load_workbook
import csv
import cv2
import io
import os
import numpy as np
from pytesseract import*
from pytesseract import image_to_string, pytesseract

# Explicitly set Tesseract executable path
pytesseract.tesseract_cmd = r'C:/Program Files/Tesseract-OCR/tesseract.exe'



class QuestionRotation:
    def __init__(self):
        self.BookSectionCropId = 0
        self.PublisherId = 0
        self.AnswersOptionsCount = 0
        self.ImageUrl = ""
        
        self.A_X1_Coordinate = 0
        self.A_Y1_Coordinate = 0

        self.B_X1_Coordinate = 0
        self.B_Y1_Coordinate = 0

        self.C_X1_Coordinate = 0
        self.C_Y1_Coordinate = 0

        self.D_X1_Coordinate = 0
        self.D_Y1_Coordinate = 0

        self.E_X1_Coordinate = 0
        self.E_Y1_Coordinate = 0

        self.A_multWidth = 0
        self.A_multHeight = 0

        self.B_multWidth = 0
        self.B_multHeight = 0

        self.C_multWidth = 0
        self.C_multHeight = 0

        self.D_multWidth = 0
        self.D_multHeight = 0

        self.E_multWidth = 0
        self.E_multHeight = 0


class AllQuestions:
    def __init__(self, id, publisher_id, image_url, is_image_large):
        self.id = id
        self.publisher_id = publisher_id
        self.image_url = image_url
        self.is_image_large = is_image_large

class Program:
    path_desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    excel_path = "ocrr.xlsx"
    csv_config = {
        "has_header_record": False,
        "missing_field_found": None,
        "delimiter": ";"
    }
    header_written = False
    number = 1
    excel_dtos = [] 
   
    
    @staticmethod
    async def main():
       await Program.read_excel()

    @staticmethod
    async def read_excel():
        path = "C:/Users/senat/OneDrive/Masaüstü/ocrr.xlsx"
        excel = load_workbook(path)
        sheet = excel.active


        for row in sheet.iter_rows(min_row=4, values_only=True):
            id, publisher_id, image_url, is_image_large = row[0], row[1], row[2], row[3]

            id= int(id) if id is not None else 0
            add_List = AllQuestions(id,publisher_id,image_url,is_image_large)
            Program.excel_dtos.append(add_List)

        await Program.get_question(Program.excel_dtos)

    @staticmethod
    async def get_question(excel_dtos):
        for item in excel_dtos:
            try:
                if not item.image_url.startswith(('http://', 'https://')):
                    raise ValueError("Invalid URL")
                
                response = requests.get(item.image_url)
                response.raise_for_status() 

                image_data = Image.open(io.BytesIO(response.content))
             
                img_np = np.array(image_data) 
                gray = cv2.cvtColor(img_np, cv2.COLOR_BGR2GRAY)

                text =  pytesseract.image_to_data(gray,output_type=Output.DICT)
                #print(str_text)

                question1 = QuestionRotation()
                question1.BookSectionCropId = item.id
                question1.PublisherId = item.publisher_id
                question1.ImageUrl = item.image_url
                answersOption = 0
            
                data = []
                for i in range(0,len(text["text"])):
                        result = text["text"][i]
                        if "A)" in result:
                            print("Eşleştirildi. A:" + result)
                            x = text["left"][i] 
                            y = text["top"][i] 
                            w = text["width"][i] 
                            h = text["height"][i] 
                            question1.A_multHeight = h
                            question1.A_multWidth = w
                            question1.A_X1_Coordinate = x
                            question1.A_Y1_Coordinate = y
                            answersOption = 1
                            cv2.rectangle(img_np, (x, y), (x + w, y + h), (0, 0, 255), 2)
                        elif "B)" in result:
                            print("Eşleştirildi. B:" + result)
                            x = text["left"][i] 
                            y = text["top"][i] 
                            w = text["width"][i] 
                            h = text["height"][i] 
                            question1.B_multHeight = h
                            question1.B_multWidth = w
                            question1.B_X1_Coordinate = x
                            question1.B_Y1_Coordinate = y
                            answersOption = 2
                            cv2.rectangle(img_np, (x, y), (x + w, y + h), (0, 0, 255), 2)
                        elif "C)" in result:
                            print("Eşleştirildi. C:" + result)
                            x = text["left"][i] 
                            y = text["top"][i] 
                            w = text["width"][i] 
                            h = text["height"][i] 
                            question1.C_multHeight = h
                            question1.C_multWidth = w
                            question1.C_X1_Coordinate = x
                            question1.C_Y1_Coordinate = y
                            answersOption = 3
                            cv2.rectangle(img_np, (x, y), (x + w, y + h), (0, 0, 255), 2)
                        elif "D)" in result:
                            print("Eşleştirildi. D:" + result)
                            x = text["left"][i] 
                            y = text["top"][i] 
                            w = text["width"][i] 
                            h = text["height"][i] 
                            question1.D_multHeight = h
                            question1.D_multWidth = w
                            question1.D_X1_Coordinate = x
                            question1.D_Y1_Coordinate = y
                            answersOption = 4
                            cv2.rectangle(img_np, (x, y), (x + w, y + h), (0, 0, 255), 2)
                        elif "E)" in result:
                            print("Eşleştirildi. E:" + result)
                            x = text["left"][i] 
                            y = text["top"][i] 
                            w = text["width"][i] 
                            h = text["height"][i] 
                            question1.E_multHeight = h
                            question1.E_multWidth = w
                            question1.E_X1_Coordinate = x
                            question1.E_Y1_Coordinate = y
                            answersOption = 5
                            cv2.rectangle(img_np, (x, y), (x + w, y + h), (0, 0, 255), 2)

                question1.AnswersOptionsCount = answersOption
                
                data.append(question1)   
                print(img_np)
                print(gray)
                await Program.download_Image(img_np,str(question1.BookSectionCropId))
                cv2.imshow("Image", img_np)

                cv2.destroyAllWindows()
                
                await Program.csv_Write(data)
                
            except Exception as ex:
                print(f"Hata oluştu: {ex}")



    @staticmethod
    async def csv_Write(data):
        try:
            csv_file_path = os.path.join("C:/Users/senat/OneDrive/Masaüstü/", "QuestionCoordinates.csv")

            os.makedirs(os.path.dirname(csv_file_path), exist_ok=True)

            file_exists = os.path.isfile(csv_file_path)

            with open(csv_file_path, 'a', newline='') as csvfile:
                fieldnames = ['BookSectionCropId', 'PublisherId', 'AnswersOptionsCount', 'ImageUrl', 'A_multHeight',
                            'A_multWidth', 'A_X1_Coordinate', 'A_Y1_Coordinate', 'B_multHeight', 'B_multWidth',
                            'B_X1_Coordinate', 'B_Y1_Coordinate', 'C_multHeight', 'C_multWidth', 'C_X1_Coordinate',
                            'C_Y1_Coordinate', 'D_multHeight', 'D_multWidth', 'D_X1_Coordinate', 'D_Y1_Coordinate',
                            'E_multHeight', 'E_multWidth', 'E_X1_Coordinate', 'E_Y1_Coordinate']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

                if not file_exists:
                    writer.writeheader()
                    print("File created.")

                for item in data:
                    writer.writerow(vars(item))

        except Exception as ex:
            print(f"Hata oluştu: {ex}")
       
    @staticmethod
    async def download_Image(img,name):
        try:
                print(name)
                string_name = str(name)
                output_folder = "C:/Users/senat/OneDrive/Masaüstü/img_path"
                img_file_name = os.path.join(output_folder, f"{string_name}.jpg")

                if not os.path.exists(output_folder):
                    os.makedirs(output_folder)

                success = Image.fromarray(img).save(img_file_name)

                
                if success:
                    print(f"Dosya başarıyla kaydedildi: {img_file_name}")
                else:
                    print(f"Dosya kaydetme başarısız.")

                if os.path.isfile(img_file_name):
                    print(f"Dosya başarıyla kaydedildi: {img_file_name}")
                else:
                    print("Dosya kaydetme başarısız.")
                    
        except Exception as ex:

            print(f"Indırmede Hata oluştu: {ex}")  

if __name__ == "__main__":

    asyncio.run(Program.main())




